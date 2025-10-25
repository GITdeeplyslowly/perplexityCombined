"""
Shared dashboard components for Excel report generation.
Extracted from backtest results for reuse in forward test reports.
"""
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    
logger = logging.getLogger(__name__)

class DashboardStyleManager:
    """Manages consistent styling across all dashboard reports."""
    
    def __init__(self, scale_factor: float = 1.0):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for dashboard components")
            
        self.scale_factor = scale_factor
        self._setup_fonts()
        self._setup_fills()
        self._setup_borders()
        self._setup_alignments()
    
    def _setup_fonts(self):
        """Initialize font styles with scaling support."""
        base_size = int(12 * self.scale_factor)
        self.fonts = {
            'title': Font(size=base_size + 6, bold=True, color="FFFFFF"),
            'header': Font(size=base_size + 0, bold=True, color="FFFFFF"),
            'subheader': Font(size=base_size - 1, bold=True),
            'normal': Font(size=base_size - 2),
            'metric_label': Font(size=base_size - 2, bold=True),
            'metric_value': Font(size=base_size + 2, bold=True),
            'highlight': Font(size=base_size + 4, bold=True, color="FFFFFF"),
            'trade_data': Font(size=base_size - 3)
        }
    
    def _setup_fills(self):
        """Initialize fill patterns matching backtest styling."""
        self.fills = {
            'title': PatternFill(start_color="2E4BC6", end_color="2E4BC6", fill_type="solid"),
            'header': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'summary': PatternFill(start_color="E8F1FF", end_color="E8F1FF", fill_type="solid"),
            'positive': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'negative': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'neutral': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        }
    
    def _setup_borders(self):
        """Initialize border styles."""
        self.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
    
    def _setup_alignments(self):
        """Initialize alignment styles."""
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        self.center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

class DashboardLayoutManager:
    """Manages worksheet layout and positioning."""
    
    def __init__(self, worksheet, max_columns: int = 10, section_buffer: int = 2, scale_factor: float = 1.0):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for dashboard components")
            
        self.ws = worksheet
        self.max_columns = max_columns
        self.section_buffer = section_buffer
        self.scale_factor = scale_factor
        self.current_row = 1
        self.current_col = 2  # Start at column B
        self._setup_page_layout()
    
    def _setup_page_layout(self):
        """Configure page margins and column widths with scaling."""
        self.ws.page_margins = PageMargins(left=1.0, right=0.7, top=0.75, bottom=0.75)
        self.ws.column_dimensions['A'].width = 2  # Left margin buffer
        
        # Set default widths for content columns (B through end of usable range)
        # Base width per column, scaled by font size
        base_column_width = 12
        scaled_width = base_column_width * self.scale_factor
        
        for col_idx in range(self.current_col, self.current_col + self.max_columns):
            column_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[column_letter].width = scaled_width
    
    def get_usable_range(self) -> Tuple[int, int]:
        """Return start and end column indices for content."""
        start_col = self.current_col
        end_col = self.current_col + self.max_columns - 1
        return start_col, end_col
    
    def advance_row(self, rows: int = 1, add_section_spacing: bool = False):
        """Advance current row position with optional section spacing."""
        self.current_row += rows
        if add_section_spacing:
            self.current_row += self.section_buffer
    
    def merge_cells_range(self, start_col: int, end_col: int, rows: int = 1) -> str:
        """Merge cells and return the range string."""
        if start_col == end_col and rows == 1:
            return f"{get_column_letter(start_col)}{self.current_row}"
        
        start_cell = f"{get_column_letter(start_col)}{self.current_row}"
        end_cell = f"{get_column_letter(end_col)}{self.current_row + rows - 1}"
        self.ws.merge_cells(f"{start_cell}:{end_cell}")
        return start_cell

class DashboardTableBuilder:
    """Builds formatted tables and sections for dashboard reports."""
    
    def __init__(self, layout_manager: DashboardLayoutManager, style_manager: DashboardStyleManager):
        self.layout = layout_manager
        self.style = style_manager
    
    def create_title_section(self, title: str, subtitle: str = None):
        """Create main dashboard title section."""
        start_col, end_col = self.layout.get_usable_range()
        
        # Main title
        cell_ref = self.layout.merge_cells_range(start_col, end_col)
        title_cell = self.layout.ws[cell_ref]
        title_cell.value = title
        title_cell.font = self.style.fonts['title']
        title_cell.fill = self.style.fills['title']
        title_cell.alignment = self.style.center_align
        self.layout.ws.row_dimensions[self.layout.current_row].height = 30
        
        self.layout.advance_row(1, add_section_spacing=True)
        
        # Subtitle if provided
        if subtitle:
            cell_ref = self.layout.merge_cells_range(start_col, end_col)
            subtitle_cell = self.layout.ws[cell_ref]
            subtitle_cell.value = subtitle
            subtitle_cell.font = self.style.fonts['subheader']
            subtitle_cell.alignment = self.style.center_align
            self.layout.advance_row(1, add_section_spacing=True)
    
    def create_highlight_metric(self, label: str, value: str, is_positive: bool = None):
        """Create highlighted key metric display."""
        start_col, end_col = self.layout.get_usable_range()
        
        # Center the metric within available columns
        center_width = max(4, (end_col - start_col) // 2)
        center_start = start_col + 2
        center_end = center_start + center_width
        
        # Metric label
        cell_ref = self.layout.merge_cells_range(center_start, center_end)
        label_cell = self.layout.ws[cell_ref]
        label_cell.value = label
        label_cell.font = self.style.fonts['metric_label']
        label_cell.fill = self.style.fills['summary']
        label_cell.alignment = self.style.center_align
        label_cell.border = self.style.border
        self.layout.ws.row_dimensions[self.layout.current_row].height = 25
        self.layout.advance_row(1)
        
        # Metric value
        cell_ref = self.layout.merge_cells_range(center_start, center_end)
        value_cell = self.layout.ws[cell_ref]
        value_cell.value = value
        value_cell.font = self.style.fonts['highlight']
        value_cell.alignment = self.style.center_align
        value_cell.border = self.style.border
        
        # Apply conditional formatting
        if is_positive is not None:
            value_cell.fill = self.style.fills['positive'] if is_positive else self.style.fills['negative']
        else:
            value_cell.fill = self.style.fills['neutral']
        
        self.layout.ws.row_dimensions[self.layout.current_row].height = 35
        self.layout.advance_row(1, add_section_spacing=True)
    
    def create_metrics_table(self, metrics_data: List[Tuple[str, Any]], title: str = "PERFORMANCE SUMMARY"):
        """Create performance metrics table."""
        start_col, end_col = self.layout.get_usable_range()
        
        # Section title
        cell_ref = self.layout.merge_cells_range(start_col, end_col)
        title_cell = self.layout.ws[cell_ref]
        title_cell.value = title
        title_cell.font = self.style.fonts['header']
        title_cell.fill = self.style.fills['header']
        title_cell.alignment = self.style.center_align
        self.layout.ws.row_dimensions[self.layout.current_row].height = 25
        self.layout.advance_row(1)
        
        # Create two-column layout for metrics
        metrics_per_row = 2
        col_width = (end_col - start_col + 1) // metrics_per_row
        
        for i in range(0, len(metrics_data), metrics_per_row):
            row_metrics = metrics_data[i:i + metrics_per_row]
            
            for j, (label, value) in enumerate(row_metrics):
                # Fixed column positions for better visibility (per user requirements)
                # Left side: Label in C(3), Value in D(4)  
                # Right side: Label in H(8), Value in I(9)
                if j == 0:  # Left side metrics
                    label_col = 3  # Column C
                    value_col = 4  # Column D
                else:  # Right side metrics  
                    label_col = 8  # Column H (moved from I)
                    value_col = 9  # Column I (moved from L)
                
                # Metric label
                label_cell = self.layout.ws.cell(row=self.layout.current_row, column=label_col, value=label)
                label_cell.font = self.style.fonts['metric_label']
                label_cell.border = self.style.border
                label_cell.alignment = self.style.left_align
                
                # Metric value
                value_cell = self.layout.ws.cell(row=self.layout.current_row, column=value_col, value=value)
                value_cell.font = self.style.fonts['normal']
                value_cell.border = self.style.border
                value_cell.alignment = self.style.center_align
            
            self.layout.advance_row(1)
        
        self.layout.advance_row(0, add_section_spacing=True)
    
    def create_config_table(self, config_df: pd.DataFrame, title: str = "STRATEGY CONFIGURATION"):
        """Create configuration parameters table."""
        start_col, end_col = self.layout.get_usable_range()
        
        # Section title
        cell_ref = self.layout.merge_cells_range(start_col, end_col)
        title_cell = self.layout.ws[cell_ref]
        title_cell.value = title
        title_cell.font = self.style.fonts['header']
        title_cell.fill = self.style.fills['header']
        title_cell.alignment = self.style.center_align
        self.layout.ws.row_dimensions[self.layout.current_row].height = 25
        self.layout.advance_row(1)
        
        # Write configuration data
        for _, row in config_df.iterrows():
            # Parameter name
            param_cell = self.layout.ws.cell(row=self.layout.current_row, column=start_col, value=row.iloc[0])
            param_cell.font = self.style.fonts['metric_label']
            param_cell.border = self.style.border
            
            # Parameter value
            value_cell = self.layout.ws.cell(row=self.layout.current_row, column=start_col + 2, value=row.iloc[1])
            value_cell.font = self.style.fonts['normal']
            value_cell.border = self.style.border
            
            self.layout.advance_row(1)
        
        self.layout.advance_row(0, add_section_spacing=True)
    
    def create_trades_table(self, trades_df: pd.DataFrame, title: str = "DETAILED TRADES LOG"):
        """Create detailed trades table using DataFrame direct write."""
        if trades_df.empty:
            self._create_no_trades_message(title)
            return
        
        # Debug logging
        logger.info(f"Creating trades table with {len(trades_df)} rows, {len(trades_df.columns)} columns")
        logger.info(f"Trades DataFrame columns: {list(trades_df.columns)}")
        logger.info(f"Trades DataFrame shape: {trades_df.shape}")
        
        start_col, end_col = self.layout.get_usable_range()
        
        # Section title
        cell_ref = self.layout.merge_cells_range(start_col, end_col)
        title_cell = self.layout.ws[cell_ref]
        title_cell.value = title
        title_cell.font = self.style.fonts['header']
        title_cell.fill = self.style.fills['header']
        title_cell.alignment = self.style.center_align
        self.layout.ws.row_dimensions[self.layout.current_row].height = 25
        self.layout.advance_row(1)
        
        # Write DataFrame directly using dataframe_to_rows
        for r_idx, row in enumerate(dataframe_to_rows(trades_df, index=False, header=True), self.layout.current_row):
            for c_idx, value in enumerate(row, start_col):
                cell = self.layout.ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == self.layout.current_row:  # Header row
                    cell.font = self.style.fonts['header']
                    cell.fill = self.style.fills['header']
                    cell.alignment = self.style.center_wrap
                else:  # Data rows
                    cell.font = self.style.fonts['trade_data']
                    cell.alignment = self.style.center_align
                    
                    # Apply number formatting for financial columns
                    if any(currency_term in str(trades_df.columns[c_idx - start_col] if c_idx - start_col < len(trades_df.columns) else "") 
                           for currency_term in ['PnL', 'Price', 'Capital', 'Commission']):
                        if isinstance(value, (int, float)) and value != 'Open':
                            cell.number_format = "#,##0.00"
                    
                    # Color code P&L columns
                    if any(pnl_term in str(trades_df.columns[c_idx - start_col] if c_idx - start_col < len(trades_df.columns) else "")
                           for pnl_term in ['PnL', 'Gross PnL', 'Net PnL']) and isinstance(value, (int, float)):
                        if value > 0:
                            cell.fill = self.style.fills['positive']
                        elif value < 0:
                            cell.fill = self.style.fills['negative']
                
                cell.border = self.style.border
        
        # Update current row position
        self.layout.current_row = self.layout.current_row + len(trades_df) + 1  # +1 for header
        self.layout.advance_row(0, add_section_spacing=True)
        
        # Set column widths for better display
        self._set_trades_column_widths(start_col, len(trades_df.columns))
    
    def _create_no_trades_message(self, title: str):
        """Create message when no trades are available."""
        start_col, end_col = self.layout.get_usable_range()
        
        # Section title
        cell_ref = self.layout.merge_cells_range(start_col, end_col)
        title_cell = self.layout.ws[cell_ref]
        title_cell.value = title
        title_cell.font = self.style.fonts['header']
        title_cell.fill = self.style.fills['header']
        title_cell.alignment = self.style.center_align
        self.layout.advance_row(1)
        
        # No trades message
        cell_ref = self.layout.merge_cells_range(start_col, end_col)
        message_cell = self.layout.ws[cell_ref]
        message_cell.value = "No trades executed during this forward test period"
        message_cell.font = self.style.fonts['normal']
        message_cell.alignment = self.style.center_align
        self.layout.advance_row(1, add_section_spacing=True)
    
    def _set_trades_column_widths(self, start_col: int, num_columns: int):
        """Set appropriate column widths for trades table, scaled for font size."""
        # Base column widths for common trade table columns (updated for 11 columns)
        # [#, Entry Time, Exit Time, Entry Price, Exit Price, Qty, Gross PnL, Commission, Net PnL, Exit Reason, Duration (min)]
        base_widths = [4, 18, 18, 12, 12, 8, 13, 12, 11, 15, 12]
        
        # Scale widths based on font scale factor
        scale_factor = self.style.scale_factor
        
        for i in range(num_columns):
            col_idx = start_col + i
            base_width = base_widths[i] if i < len(base_widths) else 12
            scaled_width = base_width * scale_factor
            column_letter = get_column_letter(col_idx)
            self.layout.ws.column_dimensions[column_letter].width = scaled_width