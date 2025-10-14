#!/usr/bin/env python3
"""
Test the merged cell fix for forward_test_results.py
"""

def test_merged_cell_fix():
    """Test that merged cell creation works without errors"""
    print("Testing merged cell fix...")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        
        # Create test workbook
        wb = Workbook()
        ws = wb.active
        
        # Test the exact same approach as in the code
        config_text = "TEST CONFIG\nLine 1\nLine 2\nLine 3"
        start_row, start_col = 15, 2
        end_row, end_col = start_row + 8, start_col + 4
        
        # Set value and formatting BEFORE merging (the fix)
        config_cell = ws.cell(row=start_row, column=start_col)
        config_cell.value = f"STRATEGY CONFIGURATION\n\n{config_text}"
        config_cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Now merge the cells
        cell_range = f"{config_cell.coordinate}:{ws.cell(row=end_row, column=end_col).coordinate}"
        ws.merge_cells(cell_range)
        
        # Test saving
        wb.save("test_merged_cell_fix.xlsx")
        
        print("✅ SUCCESS: Merged cell creation works correctly")
        print("✅ Fixed: Set value BEFORE merging cells")
        print("✅ No 'MergedCell' object attribute 'value' is read-only error")
        
        return True
        
    except Exception as e:
        print(f"❌ FAILED: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_merged_cell_fix()
    if success:
        print("\n🎯 Merged cell fix verified - forward test should work now!")
    else:
        print("\n💥 Merged cell fix failed - need different approach")