"""
backtest/results.py

Compute trading performance metrics for backtests, including:
- Total returns
- Win/loss stats
- Profit factor
- Equity curve and drawdown
- Trade durations and reason summaries
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple
from dataclasses import dataclass
from utils.time_utils import format_timestamp
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Optional helpers (safe_divide and drawdown logic)
def safe_divide(numerator, denominator, default=0.0):
    try:
        if denominator == 0:
            return default
        return numerator / denominator
    except:
        return default

def calculate_drawdown(equity_curve: List[float]) -> float:
    max_drawdown = 0.0
    peak = equity_curve[0] if equity_curve else 0
    for value in equity_curve:
        if value > peak:
            peak = value
        drawdown = (peak - value) / peak if peak > 0 else 0
        max_drawdown = max(max_drawdown, drawdown)
    return max_drawdown * 100


@dataclass
class TradeResult:
    """Single structured record for a completed trade."""
    entry_time: datetime
    exit_time: datetime
    entry_price: float
    exit_price: float
    quantity: int
    pnl: float
    commission: float
    exit_reason: str


@dataclass
class TradingMetrics:
    """Aggregate statistics for evaluation."""
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    gross_profit: float = 0.0
    gross_loss: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    total_pnl: float = 0.0
    total_commission: float = 0.0
    net_pnl: float = 0.0
    best_trade: float = 0.0
    worst_trade: float = 0.0
    return_percent: float = 0.0
    final_capital: float = 0.0
    profit_factor: float = 0.0
    drawdown_percent: float = 0.0


class Results:
    """
    Result engine. Stores completed trades, tracks equity progression,
    and outputs performance metrics and CSV reports.
    """

    def __init__(self, initial_capital: float):
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.trades: List[TradeResult] = []
        self.equity_curve: List[Tuple[datetime, float]] = []

    def add_trade(self, trade_data: Dict[str, Any]) -> None:
        """Appends a trade and updates capital/equity."""
        trade = TradeResult(
            entry_time=trade_data['entry_time'],
            exit_time=trade_data['exit_time'],
            entry_price=trade_data['entry_price'],
            exit_price=trade_data['exit_price'],
            quantity=trade_data['quantity'],
            pnl=trade_data['pnl'],
            commission=trade_data['commission'],
            exit_reason=trade_data.get('exit_reason', ''),
        )
        self.trades.append(trade)
        self.current_capital += (trade.pnl - trade.commission)
        self.equity_curve.append((trade.exit_time, self.current_capital))

    def calculate_metrics(self) -> TradingMetrics:
        if not self.trades:
            return TradingMetrics(final_capital=self.initial_capital)

        total_trades = len(self.trades)
        wins = [t for t in self.trades if t.pnl > 0]
        losses = [t for t in self.trades if t.pnl < 0]

        gross_profit = sum(t.pnl for t in wins)
        gross_loss = sum(t.pnl for t in losses)
        total_commission = sum(t.commission for t in self.trades)
        total_pnl = gross_profit + gross_loss
        net_pnl = total_pnl - total_commission

        avg_win = safe_divide(gross_profit, len(wins))
        avg_loss = safe_divide(abs(gross_loss), len(losses))

        final_capital = self.initial_capital + net_pnl
        return_percent = safe_divide(net_pnl, self.initial_capital, 0.0) * 100
        drawdown = calculate_drawdown([v for _, v in self.equity_curve])

        return TradingMetrics(
            total_trades=total_trades,
            winning_trades=len(wins),
            losing_trades=len(losses),
            win_rate=100 * safe_divide(len(wins), total_trades),
            gross_profit=gross_profit,
            gross_loss=gross_loss,
            avg_win=avg_win,
            avg_loss=avg_loss,
            total_pnl=total_pnl,
            total_commission=total_commission,
            net_pnl=net_pnl,
            best_trade=max((t.pnl for t in self.trades), default=0.0),
            worst_trade=min((t.pnl for t in self.trades), default=0.0),
            return_percent=return_percent,
            final_capital=final_capital,
            profit_factor=safe_divide(gross_profit, abs(gross_loss), 0.0),
            drawdown_percent=drawdown,
        )

    def print_summary(self):
        """Show performance metrics in console."""
        m = self.calculate_metrics()
        print(f"\n{'='*60}")
        print("TRADING PERFORMANCE SUMMARY")
        print(f"{'='*60}")
        print(f"Total Trades        : {m.total_trades}")
        print(f"Win Rate (%)        : {m.win_rate:.2f}")
        print(f"Gross Profit        : ₹{m.gross_profit:.2f}")
        print(f"Gross Loss          : ₹{m.gross_loss:.2f}")
        print(f"Avg Win             : ₹{m.avg_win:.2f}")
        print(f"Avg Loss            : ₹{m.avg_loss:.2f}")
        print(f"Net P&L             : ₹{m.net_pnl:.2f}")
        print(f"Best Trade (P&L)    : ₹{m.best_trade:.2f}")
        print(f"Worst Trade (P&L)   : ₹{m.worst_trade:.2f}")
        print(f"Return (%)          : {m.return_percent:.2f}")
        print(f"Drawdown (%)        : {m.drawdown_percent:.2f}")
        print(f"Final Capital       : ₹{m.final_capital:,.2f}")
        print(f"Profit Factor       : {m.profit_factor:.2f}")
        print(f"Total Commission    : ₹{m.total_commission:.2f}")
        print(f"{'='*60}")

    def get_trade_summary(self) -> pd.DataFrame:
        rows = []
        capital = self.initial_capital

        # Insert starting capital row
        rows.append({
            "Entry Time": "",
            "Exit Time": "",
            "Entry Price": "",
            "Exit Price": "",
            "Lots": "",
            "Total Qty": "",
            "Gross P&L": "",
            "Commission": "",
            "Net P&L": "",
            "Exit Reason": "Starting Capital",
            "Duration (min)": "",
            "Capital Outstanding": round(capital, 2)
        })

        for t in self.trades:
            net_pnl = t.pnl - t.commission
            capital += net_pnl
            lots_display = getattr(t, 'lots_traded', t.quantity // getattr(t, 'lot_size', 1)) if hasattr(t, 'lot_size') else 'N/A'
            rows.append({
                "Entry Time": t.entry_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Exit Time": t.exit_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Entry Price": round(t.entry_price, 2),
                "Exit Price": round(t.exit_price, 2),
                "Lots": lots_display,
                "Total Qty": t.quantity,
                "Gross P&L": round(t.pnl, 2),
                "Commission": round(t.commission, 2),
                "Net P&L": round(net_pnl, 2),
                "Exit Reason": t.exit_reason,
                "Duration (min)": round((t.exit_time - t.entry_time).total_seconds() / 60, 2),
                "Capital Outstanding": round(capital, 2)
            })
        return pd.DataFrame(rows)

    def get_equity_curve(self) -> pd.DataFrame:
        """Return equity curve as DataFrame."""
        if not self.equity_curve:
            return pd.DataFrame()
        return pd.DataFrame(self.equity_curve, columns=["timestamp", "capital"])

    def export_to_csv(self, output_dir: str = "results") -> str:
        """Dump trades and performance to CSV. (No separate equity file)"""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        timestamp = format_timestamp(datetime.now())

        trades_file = os.path.join(output_dir, f"trades_{timestamp}.csv")

        trades_df.to_csv(trades_file, index=False)

        return trades_file

    def export_to_excel(self, output_dir: str = "results") -> str:
        """Export trades to Excel with colored rows based on Gross P&L."""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        timestamp = format_timestamp(datetime.now())
        excel_file = os.path.join(output_dir, f"trades_{timestamp}.xlsx")
        trades_df.to_excel(excel_file, index=False)

        # Apply color formatting
        wb = load_workbook(excel_file)
        ws = wb.active
        gross_pnl_col = None
        # Find the Gross P&L column index (1-based)
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Gross P&L":
                gross_pnl_col = idx
                break

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
            cell = row[gross_pnl_col - 1]
            try:
                value = float(cell.value)
                fill = green_fill if value > 0 else red_fill if value < 0 else None
                if fill:
                    for c in row:
                        c.fill = fill
            except Exception:
                continue

        wb.save(excel_file)
        return excel_file

BacktestResults = Results
