"""
backtest/results.py

Compute trading performance metrics for backtests, including:
- Total returns
- Win/loss stats
- Profit factor
- Equity curve and drawdown
- Trade durations and reason summaries
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass
from utils.time_utils import format_timestamp
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging

# Optional helpers (safe_divide and drawdown logic)
def safe_divide(numerator, denominator, default=0.0):
    try:
        if denominator == 0:
            return default
        return numerator / denominator
    except:
        return default

def calculate_drawdown(equity_curve: List[float]) -> float:
    max_drawdown = 0.0
    peak = equity_curve[0] if equity_curve else 0
    for value in equity_curve:
        if value > peak:
            peak = value
        drawdown = (peak - value) / peak if peak > 0 else 0
        max_drawdown = max(max_drawdown, drawdown)
    return max_drawdown * 100


@dataclass
class TradeResult:
    """Single structured record for a completed trade."""
    entry_time: datetime
    exit_time: datetime
    entry_price: float
    exit_price: float
    quantity: int
    pnl: float
    commission: float
    exit_reason: str


@dataclass
class TradingMetrics:
    """Aggregate statistics for evaluation."""
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    gross_profit: float = 0.0
    gross_loss: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    total_pnl: float = 0.0
    total_commission: float = 0.0
    net_pnl: float = 0.0
    best_trade: float = 0.0
    worst_trade: float = 0.0
    return_percent: float = 0.0
    final_capital: float = 0.0
    profit_factor: float = 0.0
    drawdown_percent: float = 0.0


class Results:
    """
    Result engine. Stores completed trades, tracks equity progression,
    and outputs performance metrics and CSV reports.
    """

    def __init__(self, initial_capital: float):
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.trades: List[TradeResult] = []
        self.equity_curve: List[Tuple[datetime, float]] = []
        self.config: Optional[Dict[str, Any]] = None  # Add config attribute

    def add_trade(self, trade_data: Dict[str, Any]) -> None:
        """Appends a trade and updates capital/equity."""
        trade = TradeResult(
            entry_time=trade_data['entry_time'],
            exit_time=trade_data['exit_time'],
            entry_price=trade_data['entry_price'],
            exit_price=trade_data['exit_price'],
            quantity=trade_data['quantity'],
            pnl=trade_data['pnl'],
            commission=trade_data['commission'],
            exit_reason=trade_data.get('exit_reason', ''),
        )
        self.trades.append(trade)
        self.current_capital += (trade.pnl - trade.commission)
        self.equity_curve.append((trade.exit_time, self.current_capital))

    def set_config(self, config: Dict[str, Any]):
        self.config = config

    def _create_additional_info_table(self) -> pd.DataFrame:
        """Create additional info table with indicators and parameters"""
        if not hasattr(self, 'config') or not self.config:
            return pd.DataFrame([{"Key": "Configuration", "Value": "Not Available"}])

        rows = []
        config = self.config
        strategy_config = config.get('strategy', config)  # fallback if not nested
        risk_config = config.get('risk', config)  # fallback if not nested

        # Indicators activated
        indicator_map = {
            'use_ema_crossover': 'EMA Crossover',
            'use_macd': 'MACD',
            'use_vwap': 'VWAP',
            'use_rsi_filter': 'RSI Filter',
            'use_htf_trend': 'HTF Trend',
            'use_bollinger_bands': 'Bollinger Bands',
            'use_stochastic': 'Stochastic',
            'use_atr': 'ATR'
        }
        active_indicators = [name for key, name in indicator_map.items()
                             if strategy_config.get(key, False)]
        rows.append({"Key": "Indicators Activated",
                     "Value": ", ".join(active_indicators) if active_indicators else "None"})

        # EMA parameters
        if strategy_config.get('use_ema_crossover', False):
            ema_params = f"Fast EMA: {strategy_config.get('fast_ema', 9)}, Slow EMA: {strategy_config.get('slow_ema', 21)}"
            rows.append({"Key": "EMA Parameters", "Value": ema_params})

        # MACD parameters
        if strategy_config.get('use_macd', False):
            macd_params = f"Fast: {strategy_config.get('macd_fast', 12)}, Slow: {strategy_config.get('macd_slow', 26)}, Signal: {strategy_config.get('macd_signal', 9)}"
            rows.append({"Key": "MACD Parameters", "Value": macd_params})

        # RSI parameters
        if strategy_config.get('use_rsi_filter', False):
            rsi_params = f"Period: {strategy_config.get('rsi_length', 14)}, Overbought: {strategy_config.get('rsi_overbought', 70)}, Oversold: {strategy_config.get('rsi_oversold', 30)}"
            rows.append({"Key": "RSI Parameters", "Value": rsi_params})

        # HTF Trend parameters
        if strategy_config.get('use_htf_trend', False):
            htf_params = f"HTF Period: {strategy_config.get('htf_period', 20)}"
            rows.append({"Key": "HTF Trend Parameters", "Value": htf_params})

        # SL points
        rows.append({"Key": "SL Points", "Value": str(risk_config.get('base_sl_points', 15))})

        # TP points
        tp_points = risk_config.get('tp_points', config.get('tp_points', [10, 25, 50, 100]))
        rows.append({"Key": "TP Points", "Value": ", ".join(map(str, tp_points))})

        # Trail SL info
        trail_enabled = risk_config.get('use_trail_stop', config.get('use_trail_stop', True))
        trail_activation = risk_config.get('trail_activation_points', config.get('trail_activation_points', 25))
        trail_distance = risk_config.get('trail_distance_points', config.get('trail_distance_points', 10))
        trail_info = f"Enabled: {trail_enabled}, Activation: {trail_activation} points, Distance: {trail_distance} points"
        rows.append({"Key": "Trailing Stop", "Value": trail_info})

        # --- LOGGING FOR DEBUG ---
        green_bars_req = strategy_config.get('consecutive_green_bars')
        logging.warning(f"DEBUG: consecutive_green_bars in strategy_config: {green_bars_req}")
        logging.warning(f"DEBUG: strategy_config keys: {list(strategy_config.keys())}")
        logging.warning(f"DEBUG: config keys: {list(config.keys())}")

        rows.append({"Key": "Green Bars Required for Entry", "Value": str(green_bars_req) })

        return pd.DataFrame(rows)

    def calculate_metrics(self) -> TradingMetrics:
        if not self.trades:
            return TradingMetrics(final_capital=self.initial_capital)

        total_trades = len(self.trades)
        wins = [t for t in self.trades if t.pnl > 0]
        losses = [t for t in self.trades if t.pnl < 0]

        gross_profit = sum(t.pnl for t in wins)
        gross_loss = sum(t.pnl for t in losses)
        total_commission = sum(t.commission for t in self.trades)
        total_pnl = gross_profit + gross_loss
        net_pnl = total_pnl - total_commission

        avg_win = safe_divide(gross_profit, len(wins))
        avg_loss = safe_divide(abs(gross_loss), len(losses))

        final_capital = self.initial_capital + net_pnl
        return_percent = safe_divide(net_pnl, self.initial_capital, 0.0) * 100
        drawdown = calculate_drawdown([v for _, v in self.equity_curve])

        return TradingMetrics(
            total_trades=total_trades,
            winning_trades=len(wins),
            losing_trades=len(losses),
            win_rate=100 * safe_divide(len(wins), total_trades),
            gross_profit=gross_profit,
            gross_loss=gross_loss,
            avg_win=avg_win,
            avg_loss=avg_loss,
            total_pnl=total_pnl,
            total_commission=total_commission,
            net_pnl=net_pnl,
            best_trade=max((t.pnl for t in self.trades), default=0.0),
            worst_trade=min((t.pnl for t in self.trades), default=0.0),
            return_percent=return_percent,
            final_capital=final_capital,
            profit_factor=safe_divide(gross_profit, abs(gross_loss), 0.0),
            drawdown_percent=drawdown,
        )

    def print_summary(self):
        """Show performance metrics in console."""
        m = self.calculate_metrics()
        print(f"\n{'='*60}")
        print("TRADING PERFORMANCE SUMMARY")
        print(f"{'='*60}")
        print(f"Total Trades        : {m.total_trades}")
        print(f"Win Rate (%)        : {m.win_rate:.2f}")
        print(f"Gross Profit        : ₹{m.gross_profit:.2f}")
        print(f"Gross Loss          : ₹{m.gross_loss:.2f}")
        print(f"Avg Win             : ₹{m.avg_win:.2f}")
        print(f"Avg Loss            : ₹{m.avg_loss:.2f}")
        print(f"Net P&L             : ₹{m.net_pnl:.2f}")
        print(f"Best Trade (P&L)    : ₹{m.best_trade:.2f}")
        print(f"Worst Trade (P&L)   : ₹{m.worst_trade:.2f}")
        print(f"Return (%)          : {m.return_percent:.2f}")
        print(f"Drawdown (%)        : {m.drawdown_percent:.2f}")
        print(f"Final Capital       : ₹{m.final_capital:,.2f}")
        print(f"Profit Factor       : {m.profit_factor:.2f}")
        print(f"Total Commission    : ₹{m.total_commission:.2f}")
        print(f"{'='*60}")

    def get_trade_summary(self) -> pd.DataFrame:
        rows = []
        capital = self.initial_capital

        # Insert starting capital row
        rows.append({
            "Entry Time": "",
            "Exit Time": "",
            "Entry Price": "",
            "Exit Price": "",
            "Lots": "",
            "Total Qty": "",
            "Gross P&L": "",
            "Commission": "",
            "Net P&L": "",
            "Exit Reason": "Starting Capital",
            "Duration (min)": "",
            "Capital Outstanding": round(capital, 2)
        })

        for t in self.trades:
            net_pnl = t.pnl - t.commission
            capital += net_pnl
            lots_display = getattr(t, 'lots_traded', t.quantity // getattr(t, 'lot_size', 1)) if hasattr(t, 'lot_size') else 'N/A'
            rows.append({
                "Entry Time": t.entry_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Exit Time": t.exit_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Entry Price": round(t.entry_price, 2),
                "Exit Price": round(t.exit_price, 2),
                "Lots": lots_display,
                "Total Qty": t.quantity,
                "Gross P&L": round(t.pnl, 2),
                "Commission": round(t.commission, 2),
                "Net P&L": round(net_pnl, 2),
                "Exit Reason": t.exit_reason,
                "Duration (min)": round((t.exit_time - t.entry_time).total_seconds() / 60, 2),
                "Capital Outstanding": round(capital, 2)
            })
        return pd.DataFrame(rows)

    def get_equity_curve(self) -> pd.DataFrame:
        """Return equity curve as DataFrame."""
        if not self.equity_curve:
            return pd.DataFrame()
        return pd.DataFrame(self.equity_curve, columns=["timestamp", "capital"])

    def export_to_csv(self, output_dir: str = "results") -> str:
        """Dump trades and performance to CSV. (No separate equity file)"""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        timestamp = format_timestamp(datetime.now())
        trades_file = os.path.join(output_dir, f"trades_{timestamp}.csv")

        # Create additional info table with trading configuration
        additional_info_df = self._create_additional_info_table()

        # Write both tables to the same CSV file
        with open(trades_file, 'w', newline='') as f:
            additional_info_df.to_csv(f, index=False)
            f.write('\n')  # Add empty line between tables
            trades_df.to_csv(f, index=False)

        return trades_file

    def export_to_excel(self, output_dir: str = "results") -> str:
        """Export trades and config info to Excel with no currency formatting (currency agnostic)."""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        additional_info_df = self._create_additional_info_table()
        timestamp = format_timestamp(datetime.now())
        excel_file = os.path.join(output_dir, f"trades_{timestamp}.xlsx")

        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            additional_info_df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=0)
            trades_df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=len(additional_info_df) + 2)

        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_file)
        ws = wb.active

        header_row = len(additional_info_df) + 3

        # Style the header row: bold and center aligned
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        for col in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.value:
                    try:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                    except:
                        pass
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)

        # Identify numeric/percentage columns
        num_cols = []
        pct_cols = []

        headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
        for idx, header in enumerate(headers, start=1):
            if header and isinstance(header, str):
                h = header.lower()
                if any(keyword in h for keyword in ["p&l", "gross", "net", "commission", "capital", "price"]):
                    num_cols.append(idx)
                if any(keyword in h for keyword in ["rate", "return", "drawdown"]):
                    pct_cols.append(idx)

        # Format all relevant numeric columns without currency symbol
        for row in range(header_row + 1, ws.max_row + 1):
            for col in num_cols:
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value not in (None, ""):
                        cell.number_format = "#,##0.00"
                except:
                    pass

        # Percentage formatting
        for row in range(header_row + 1, ws.max_row + 1):
            for col in pct_cols:
                cell = ws.cell(row=row, column=col)
                try:
                    if cell.value not in (None, ""):
                        val = float(cell.value)
                        # Convert whole numbers to percentage decimal if necessary
                        if val > 1:
                            cell.value = val / 100
                        cell.number_format = "0.00%"
                except:
                    pass

        # Color code Gross P&L
        gross_pnl_col = None
        for idx, cell in enumerate(ws[header_row], 1):
            if cell.value == "Gross P&L":
                gross_pnl_col = idx
                break

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in ws.iter_rows(min_row=header_row + 1, max_col=ws.max_column):
            cell = row[gross_pnl_col - 1] if gross_pnl_col else None
            try:
                value = float(cell.value) if cell else None
                fill = green_fill if value and value > 0 else red_fill if value and value < 0 else None
                if fill:
                    for c in row:
                        c.fill = fill
            except:
                continue

        wb.save(excel_file)
        return excel_file

BacktestResults = Results
