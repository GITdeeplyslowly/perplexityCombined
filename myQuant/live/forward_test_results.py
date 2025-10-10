"""
Forward Test Results Export Module

Provides comprehensive results file generation for forward tests,
similar to the backtest module but adapted for live/forward testing.

Features:
- Configuration capture and export
- Performance metrics summary  
- All trades detailed listing
- CSV and Excel export formats
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any, List
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export will fallback to CSV")

class ForwardTestResults:
    """Forward test results export and reporting"""
    
    def __init__(self, config: Dict[str, Any], position_manager, start_time: datetime):
        self.config = config
        self.position_manager = position_manager
        self.start_time = start_time
        self.end_time = None
        
    def finalize(self):
        """Mark the forward test as completed"""
        self.end_time = datetime.now()
        
    def export_to_csv(self, output_dir: str = "results") -> str:
        """Export forward test results to CSV format"""
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"ForwardTest_Results_{timestamp}.csv")
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            # Write header
            f.write("FORWARD TEST RESULTS REPORT\\n")
            f.write("=" * 80 + "\\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\\n")
            f.write(f"Test Duration: {self._get_test_duration()}\\n")
            f.write("\\n")
            
            # Configuration Section
            f.write("CONFIGURATION\\n")
            f.write("-" * 40 + "\\n")
            f.write(self._generate_config_text())
            f.write("\\n\\n")
            
            # Performance Summary
            f.write("PERFORMANCE SUMMARY\\n")
            f.write("-" * 40 + "\\n")
            f.write(self._generate_performance_text())
            f.write("\\n\\n")
            
            # Trades Section
            f.write("ALL TRADES DETAIL\\n")
            f.write("-" * 40 + "\\n")
            
        # Append trades data as CSV table
        trades_df = self._get_trades_dataframe()
        if not trades_df.empty:
            trades_df.to_csv(filename, mode='a', index=False)
        else:
            with open(filename, 'a') as f:
                f.write("No trades executed during this forward test\\n")
                
        logger.info(f"Forward test results exported to: {filename}")
        return filename
        
    def export_to_excel(self, output_dir: str = "results") -> str:
        """Export forward test results to Excel format"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("Excel export not available, falling back to CSV")
            return self.export_to_csv(output_dir)
            
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"ForwardTest_Results_{timestamp}.xlsx")
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary)
        
        # Configuration Sheet  
        ws_config = wb.create_sheet("Configuration")
        self._create_config_sheet(ws_config)
        
        # Trades Sheet
        ws_trades = wb.create_sheet("Trades")
        self._create_trades_sheet(ws_trades)
        
        wb.save(filename)
        logger.info(f"Forward test Excel results exported to: {filename}")
        return filename
        
    def _get_test_duration(self) -> str:
        """Get formatted test duration"""
        if not self.end_time:
            return "In Progress"
        duration = self.end_time - self.start_time
        hours, remainder = divmod(int(duration.total_seconds()), 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
    def _generate_config_text(self) -> str:
        """Generate configuration section as text"""
        lines = []
        
        # Instrument
        inst = self.config.get('instrument', {})
        lines.append(f"Symbol: {inst.get('symbol', 'N/A')}")
        lines.append(f"Exchange: {inst.get('exchange', 'N/A')}")
        lines.append(f"Product Type: {inst.get('product_type', 'N/A')}")
        lines.append(f"Lot Size: {inst.get('lot_size', 'N/A')}")
        lines.append("")
        
        # Capital & Risk
        capital = self.config.get('capital', {})
        risk = self.config.get('risk', {})
        lines.append(f"Initial Capital: ₹{capital.get('initial_capital', 0):,.2f}")
        lines.append(f"Max Trades/Day: {risk.get('max_positions_per_day', 'N/A')}")
        lines.append(f"Risk per Trade: {risk.get('risk_per_trade_percent', 0)}%")
        lines.append("")
        
        # Data Source
        data_sim = self.config.get('data_simulation', {})
        if data_sim.get('enabled', False):
            lines.append(f"Data Source: File Simulation")
            lines.append(f"File: {data_sim.get('file_path', 'N/A')}")
        else:
            lines.append(f"Data Source: Live WebStream")
        lines.append("")
        
        return "\\n".join(lines)
        
    def _generate_performance_text(self) -> str:
        """Generate performance summary as text"""
        perf = self.position_manager.get_performance_summary()
        
        lines = []
        lines.append(f"Total Trades: {perf['total_trades']}")
        lines.append(f"Winning Trades: {perf['winning_trades']}")
        lines.append(f"Losing Trades: {perf['losing_trades']}")
        lines.append(f"Win Rate: {perf['win_rate']:.1f}%")
        lines.append(f"Total P&L: ₹{perf['total_pnl']:,.2f}")
        lines.append(f"Profit Factor: {perf['profit_factor']:.2f}")
        lines.append("")
        
        # Capital performance
        initial = self.position_manager.initial_capital
        current = self.position_manager.available_capital
        change = current - initial
        change_pct = (change / initial) * 100 if initial > 0 else 0
        
        lines.append(f"Initial Capital: ₹{initial:,.2f}")
        lines.append(f"Final Capital: ₹{current:,.2f}")
        lines.append(f"Capital Change: ₹{change:,.2f} ({change_pct:+.2f}%)")
        
        return "\\n".join(lines)
        
    def _get_trades_dataframe(self) -> pd.DataFrame:
        """Get all trades as DataFrame"""
        if not self.position_manager.completed_trades:
            return pd.DataFrame()
            
        trades_data = []
        for i, trade in enumerate(self.position_manager.completed_trades, 1):
            trades_data.append({
                'Trade #': i,
                'Entry Time': trade.entry_time.strftime('%Y-%m-%d %H:%M:%S'),
                'Exit Time': trade.exit_time.strftime('%Y-%m-%d %H:%M:%S') if trade.exit_time else 'Open',
                'Side': trade.side,
                'Quantity': trade.quantity,
                'Entry Price': f"₹{trade.entry_price:.2f}",
                'Exit Price': f"₹{trade.exit_price:.2f}" if trade.exit_price else 'N/A',
                'Gross P&L': f"₹{trade.gross_pnl:.2f}",
                'Commission': f"₹{trade.commission:.2f}",
                'Net P&L': f"₹{trade.net_pnl:.2f}",
                'Exit Reason': trade.exit_reason or 'N/A'
            })
            
        return pd.DataFrame(trades_data)
        
    def _create_summary_sheet(self, ws):
        """Create Excel summary sheet"""
        row = 1
        
        # Title
        ws.cell(row=row, column=1, value="FORWARD TEST RESULTS").font = Font(size=16, bold=True)
        row += 2
        
        # Test Info
        ws.cell(row=row, column=1, value="Test Duration:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=self._get_test_duration())
        row += 1
        
        ws.cell(row=row, column=1, value="Generated:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 2
        
        # Performance metrics
        perf = self.position_manager.get_performance_summary()
        metrics = [
            ("Total Trades", perf['total_trades']),
            ("Win Rate", f"{perf['win_rate']:.1f}%"),
            ("Total P&L", f"₹{perf['total_pnl']:,.2f}"),
            ("Profit Factor", f"{perf['profit_factor']:.2f}"),
        ]
        
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
    def _create_config_sheet(self, ws):
        """Create Excel configuration sheet"""
        config_text = self._generate_config_text()
        lines = config_text.split('\\n')
        
        for i, line in enumerate(lines, 1):
            ws.cell(row=i, column=1, value=line)
            
    def _create_trades_sheet(self, ws):
        """Create Excel trades sheet"""
        trades_df = self._get_trades_dataframe()
        
        if trades_df.empty:
            ws.cell(row=1, column=1, value="No trades executed during this forward test")
            return
            
        # Add headers and data
        for r_idx, row in enumerate(dataframe_to_rows(trades_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row
                    cell.font = Font(bold=True)
                    
    def get_results_summary(self) -> Dict[str, Any]:
        """Get complete results summary for GUI display"""
        return {
            'config': self.config,
            'performance': self.position_manager.get_performance_summary(),
            'trades': self._get_trades_dataframe(),
            'duration': self._get_test_duration(),
            'start_time': self.start_time,
            'end_time': self.end_time
        }