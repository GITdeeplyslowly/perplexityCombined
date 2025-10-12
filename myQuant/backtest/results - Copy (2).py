"""
backtest/results.py

Compute trading performance metrics for backtests, including:
- Total returns
- Win/loss stats
- Profit factor
- Equity curve and drawdown
- Trade durations and reason summaries
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass
from utils.time_utils import format_timestamp
import os
from openpyxl import load_workbook, Workbook  # type: ignore[reportMissingModuleSource]
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # type: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter  # type: ignore[reportMissingModuleSource]
from openpyxl.worksheet.page import PageMargins  # type: ignore[reportMissingModuleSource]
import logging

# Optional helpers (safe_divide and drawdown logic)
def safe_divide(numerator, denominator, default=0.0):
    try:
        if denominator == 0:
            return default
        return numerator / denominator
    except:
        return default

def calculate_drawdown(equity_curve: List[float]) -> float:
    max_drawdown = 0.0
    peak = equity_curve[0] if equity_curve else 0
    for value in equity_curve:
        if value > peak:
            peak = value
        drawdown = (peak - value) / peak if peak > 0 else 0
        max_drawdown = max(max_drawdown, drawdown)
    return max_drawdown * 100


@dataclass
class TradeResult:
    """Single structured record for a completed trade."""
    entry_time: datetime
    exit_time: datetime
    entry_price: float
    exit_price: float
    quantity: int
    pnl: float
    commission: float
    exit_reason: str


@dataclass
class TradingMetrics:
    """Aggregate statistics for evaluation."""
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    gross_profit: float = 0.0
    gross_loss: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    total_pnl: float = 0.0
    total_commission: float = 0.0
    net_pnl: float = 0.0
    best_trade: float = 0.0
    worst_trade: float = 0.0
    return_percent: float = 0.0
    final_capital: float = 0.0
    profit_factor: float = 0.0
    drawdown_percent: float = 0.0


class Results:
    """
    Result engine. Stores completed trades, tracks equity progression,
    and outputs performance metrics and CSV reports.
    """

    def __init__(self, initial_capital: float):
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.trades: List[TradeResult] = []
        self.equity_curve: List[Tuple[datetime, float]] = []
        self.config: Optional[Dict[str, Any]] = None  # Add config attribute

    def add_trade(self, trade_data: Dict[str, Any]) -> None:
        """Appends a trade and updates capital/equity."""
        trade = TradeResult(
            entry_time=trade_data['entry_time'],
            exit_time=trade_data['exit_time'],
            entry_price=trade_data['entry_price'],
            exit_price=trade_data['exit_price'],
            quantity=trade_data['quantity'],
            pnl=trade_data['pnl'],
            commission=trade_data['commission'],
            exit_reason=trade_data.get('exit_reason', ''),
        )
        self.trades.append(trade)
        self.current_capital += (trade.pnl - trade.commission)
        self.equity_curve.append((trade.exit_time, self.current_capital))

    def set_config(self, config: Dict[str, Any]):
        self.config = config

    def _create_additional_info_table(self) -> pd.DataFrame:
        """Create additional info table with indicators and parameters"""
        if not hasattr(self, 'config') or not self.config:
            return pd.DataFrame([{"Key": "Configuration", "Value": "Not Available"}])

        rows = []
        config = self.config
        strategy_config = config.get('strategy', config)  # fallback if not nested
        risk_config = config.get('risk', config)  # fallback if not nested

        # Indicators activated
        indicator_map = {
            'use_ema_crossover': 'EMA Crossover',
            'use_macd': 'MACD',
            'use_vwap': 'VWAP',
            'use_rsi_filter': 'RSI Filter',
            'use_htf_trend': 'HTF Trend',
            'use_bollinger_bands': 'Bollinger Bands',
            'use_stochastic': 'Stochastic',
            'use_atr': 'ATR'
        }
        active_indicators = [name for key, name in indicator_map.items()
                             if strategy_config.get(key, False)]
        rows.append({"Key": "Indicators Activated",
                     "Value": ", ".join(active_indicators) if active_indicators else "None"})

        # EMA parameters
        if strategy_config.get('use_ema_crossover', False):
            ema_params = f"Fast EMA: {strategy_config.get('fast_ema', 9)}, Slow EMA: {strategy_config.get('slow_ema', 21)}"
            rows.append({"Key": "EMA Parameters", "Value": ema_params})

        # MACD parameters
        if strategy_config.get('use_macd', False):
            macd_params = f"Fast: {strategy_config.get('macd_fast', 12)}, Slow: {strategy_config.get('macd_slow', 26)}, Signal: {strategy_config.get('macd_signal', 9)}"
            rows.append({"Key": "MACD Parameters", "Value": macd_params})

        # RSI parameters
        if strategy_config.get('use_rsi_filter', False):
            rsi_params = f"Period: {strategy_config.get('rsi_length', 14)}, Overbought: {strategy_config.get('rsi_overbought', 70)}, Oversold: {strategy_config.get('rsi_oversold', 30)}"
            rows.append({"Key": "RSI Parameters", "Value": rsi_params})

        # HTF Trend parameters
        if strategy_config.get('use_htf_trend', False):
            htf_params = f"HTF Period: {strategy_config.get('htf_period', 20)}"
            rows.append({"Key": "HTF Trend Parameters", "Value": htf_params})

        # SL points
        rows.append({"Key": "SL Points", "Value": str(risk_config.get('base_sl_points', 15))})

        # TP points
        tp_points = risk_config.get('tp_points', config.get('tp_points', [10, 25, 50, 100]))
        rows.append({"Key": "TP Points", "Value": ", ".join(map(str, tp_points))})

        # Trail SL info
        trail_enabled = risk_config.get('use_trail_stop', config.get('use_trail_stop', True))
        trail_activation = risk_config.get('trail_activation_points', config.get('trail_activation_points', 25))
        trail_distance = risk_config.get('trail_distance_points', config.get('trail_distance_points', 10))
        trail_info = f"Enabled: {trail_enabled}, Activation: {trail_activation} points, Distance: {trail_distance} points"
        rows.append({"Key": "Trailing Stop", "Value": trail_info})

        # --- LOGGING FOR DEBUG ---
        green_bars_req = strategy_config.get('consecutive_green_bars')
        logging.warning(f"DEBUG: consecutive_green_bars in strategy_config: {green_bars_req}")
        logging.warning(f"DEBUG: strategy_config keys: {list(strategy_config.keys())}")
        logging.warning(f"DEBUG: config keys: {list(config.keys())}")

        rows.append({"Key": "Green Bars Required for Entry", "Value": str(green_bars_req) })

        return pd.DataFrame(rows)

    def calculate_metrics(self) -> TradingMetrics:
        if not self.trades:
            return TradingMetrics(final_capital=self.initial_capital)

        total_trades = len(self.trades)
        wins = [t for t in self.trades if t.pnl > 0]
        losses = [t for t in self.trades if t.pnl < 0]

        gross_profit = sum(t.pnl for t in wins)
        gross_loss = sum(t.pnl for t in losses)
        total_commission = sum(t.commission for t in self.trades)
        total_pnl = gross_profit + gross_loss
        net_pnl = total_pnl - total_commission

        avg_win = safe_divide(gross_profit, len(wins))
        avg_loss = safe_divide(abs(gross_loss), len(losses))

        final_capital = self.initial_capital + net_pnl
        return_percent = safe_divide(net_pnl, self.initial_capital, 0.0) * 100
        drawdown = calculate_drawdown([v for _, v in self.equity_curve])

        return TradingMetrics(
            total_trades=total_trades,
            winning_trades=len(wins),
            losing_trades=len(losses),
            win_rate=100 * safe_divide(len(wins), total_trades),
            gross_profit=gross_profit,
            gross_loss=gross_loss,
            avg_win=avg_win,
            avg_loss=avg_loss,
            total_pnl=total_pnl,
            total_commission=total_commission,
            net_pnl=net_pnl,
            best_trade=max((t.pnl for t in self.trades), default=0.0),
            worst_trade=min((t.pnl for t in self.trades), default=0.0),
            return_percent=return_percent,
            final_capital=final_capital,
            profit_factor=safe_divide(gross_profit, abs(gross_loss), 0.0),
            drawdown_percent=drawdown,
        )

    def print_summary(self):
        """Show performance metrics in console."""
        m = self.calculate_metrics()
        print(f"\n{'='*60}")
        print("TRADING PERFORMANCE SUMMARY")
        print(f"{'='*60}")
        print(f"Total Trades        : {m.total_trades}")
        print(f"Win Rate (%)        : {m.win_rate:.2f}")
        print(f"Gross Profit        : â‚¹{m.gross_profit:.2f}")
        print(f"Gross Loss          : â‚¹{m.gross_loss:.2f}")
        print(f"Avg Win             : â‚¹{m.avg_win:.2f}")
        print(f"Avg Loss            : â‚¹{m.avg_loss:.2f}")
        print(f"Net P&L             : â‚¹{m.net_pnl:.2f}")
        print(f"Best Trade (P&L)    : â‚¹{m.best_trade:.2f}")
        print(f"Worst Trade (P&L)   : â‚¹{m.worst_trade:.2f}")
        print(f"Return (%)          : {m.return_percent:.2f}")
        print(f"Drawdown (%)        : {m.drawdown_percent:.2f}")
        print(f"Final Capital       : â‚¹{m.final_capital:,.2f}")
        print(f"Profit Factor       : {m.profit_factor:.2f}")
        print(f"Total Commission    : â‚¹{m.total_commission:.2f}")
        print(f"{'='*60}")

    def get_trade_summary(self) -> pd.DataFrame:
        rows = []
        capital = self.initial_capital

        # Insert starting capital row
        rows.append({
            "Entry Time": "",
            "Exit Time": "",
            "Entry Price": "",
            "Exit Price": "",
            "Lots": "",
            "Total Qty": "",
            "Gross P&L": "",
            "Commission": "",
            "Net P&L": "",
            "Exit Reason": "Starting Capital",
            "Duration (min)": "",
            "Capital Outstanding": round(capital, 2)
        })

        for t in self.trades:
            net_pnl = t.pnl - t.commission
            capital += net_pnl
            lots_display = getattr(t, 'lots_traded', t.quantity // getattr(t, 'lot_size', 1)) if hasattr(t, 'lot_size') else 'N/A'
            rows.append({
                "Entry Time": t.entry_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Exit Time": t.exit_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Entry Price": round(t.entry_price, 2),
                "Exit Price": round(t.exit_price, 2),
                "Lots": lots_display,
                "Total Qty": t.quantity,
                "Gross P&L": round(t.pnl, 2),
                "Commission": round(t.commission, 2),
                "Net P&L": round(net_pnl, 2),
                "Exit Reason": t.exit_reason,
                "Duration (min)": round((t.exit_time - t.entry_time).total_seconds() / 60, 2),
                "Capital Outstanding": round(capital, 2)
            })
        return pd.DataFrame(rows)

    def get_equity_curve(self) -> pd.DataFrame:
        """Return equity curve as DataFrame."""
        if not self.equity_curve:
            return pd.DataFrame()
        return pd.DataFrame(self.equity_curve, columns=["timestamp", "capital"])

    def export_to_csv(self, output_dir: str = "results") -> str:
        """Dump trades and performance to CSV. (No separate equity file)"""
        # Use new directory structure in Perplexity Combined folder
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # Go up to Perplexity Combined
        results_dir = os.path.join(base_dir, "results", "Back Test")
        os.makedirs(results_dir, exist_ok=True)
        
        trades_df = self.get_trade_summary()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        trades_file = os.path.join(results_dir, f"bt-{timestamp}-data.csv")

        # Create additional info table with trading configuration
        additional_info_df = self._create_additional_info_table()

        # Write both tables to the same CSV file
        with open(trades_file, 'w', newline='') as f:
            additional_info_df.to_csv(f, index=False)
            f.write('\n')  # Add empty line between tables
            trades_df.to_csv(f, index=False)

        # Log the successful save
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Backtest CSV results saved to: {trades_file}")
        print(f"✅ Backtest CSV results saved: {trades_file}")

        return trades_file

    def create_enhanced_excel_report(self, output_dir: str = "results") -> str:
        """Create a comprehensive, professional Excel dashboard report with improved layout."""
        # Use Perplexity Combined/results/Back Test directory
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        output_dir = os.path.join(project_root, "results", "Back Test")
        os.makedirs(output_dir, exist_ok=True)
        
        # Get data
        trades_df = self.get_trade_summary()
        metrics = self.calculate_metrics()
        
        # Remove starting capital row for trades processing
        trades_data = trades_df.iloc[1:].copy() if not trades_df.empty else trades_df.copy()
        starting_capital = trades_df.iloc[0]['Capital Outstanding'] if not trades_df.empty else self.initial_capital
        
        # Calculate statistics
        total_trades = len(trades_data)
        winning_trades = len(trades_data[trades_data['Gross P&L'] > 0]) if not trades_data.empty else 0
        losing_trades = len(trades_data[trades_data['Gross P&L'] < 0]) if not trades_data.empty else 0
        win_rate = (winning_trades / total_trades) * 100 if total_trades > 0 else 0
        
        best_trade = trades_data['Net P&L'].max() if not trades_data.empty else 0
        worst_trade = trades_data['Net P&L'].min() if not trades_data.empty else 0
        avg_win = trades_data[trades_data['Net P&L'] > 0]['Net P&L'].mean() if not trades_data.empty else 0
        avg_loss = trades_data[trades_data['Net P&L'] < 0]['Net P&L'].mean() if not trades_data.empty else 0
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Backtest Results"
        
        # Set page margins for better spacing
        ws.page_margins = PageMargins(left=1.0, right=0.7, top=0.75, bottom=0.75)
        ws.column_dimensions['A'].width = 2  # Empty left margin column
        
        # Define styles with better readability
        title_font = Font(size=18, bold=True, color="FFFFFF")
        header_font = Font(size=12, bold=True, color="FFFFFF")
        subheader_font = Font(size=11, bold=True)
        normal_font = Font(size=10)
        
        title_fill = PatternFill(start_color="2E4BC6", end_color="2E4BC6", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        summary_fill = PatternFill(start_color="E8F1FF", end_color="E8F1FF", fill_type="solid")
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        current_row = 1
        
        # TITLE SECTION
        ws.merge_cells(f'A{current_row}:J{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = "BACKTEST RESULTS DASHBOARD"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 30
        current_row += 2
        
        # GROSS P&L HIGHLIGHT - Centered and prominent (no emojis)
        ws.merge_cells(f'C{current_row}:H{current_row}')
        ws[f'C{current_row}'].value = "TOTAL NET P&L"
        ws[f'C{current_row}'].font = Font(size=14, bold=True)
        ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'C{current_row}'].fill = summary_fill
        ws[f'C{current_row}'].border = border
        current_row += 1

        ws.merge_cells(f'C{current_row}:H{current_row}')
        ws[f'C{current_row}'].value = f"₹{metrics.net_pnl:,.2f}"
        ws[f'C{current_row}'].font = Font(size=20, bold=True)
        ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'C{current_row}'].border = border
        if metrics.net_pnl > 0:
            ws[f'C{current_row}'].fill = positive_fill
        else:
            ws[f'C{current_row}'].fill = negative_fill
        ws.row_dimensions[current_row].height = 35
        current_row += 2

        # PERFORMANCE SUMMARY - Two rows to accommodate all metrics
        ws.merge_cells(f'B{current_row}:N{current_row}')  # Expand to include all 14 metrics
        summary_cell = ws[f'B{current_row}']
        summary_cell.value = "PERFORMANCE SUMMARY"
        summary_cell.font = header_font
        summary_cell.fill = header_fill
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # First row of 7 metric pairs (14 columns) starting B to O
        first_row_metrics = [
            "Total Trades", total_trades, "Win Rate", f"{win_rate:.2f}%", "Winning Trades", winning_trades, "Losing Trades", losing_trades,
            "Gross P&L", f"₹{metrics.total_pnl:,.2f}", "Commission", f"₹{metrics.total_commission:,.2f}", "Net P&L", f"₹{metrics.net_pnl:,.2f}", "Return %", f"{metrics.return_percent:,.2f}%"
        ]
        
        for col_idx, value in enumerate(first_row_metrics, 2):  # B=2 to O=15
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Color code based on performance
            if isinstance(value, str):
                if "₹" in str(value) and value != "N/A":
                    try:
                        numeric_val = float(str(value).replace("₹", "").replace(",", ""))
                        if numeric_val > 0:
                            cell.fill = positive_fill
                        elif numeric_val < 0:
                            cell.fill = negative_fill
                    except:
                        pass
                elif "%" in str(value):
                    try:
                        if float(str(value).replace("%", "")) > 50:
                            cell.fill = positive_fill
                        else:
                            cell.fill = neutral_fill
                    except:
                        pass
        current_row += 1

        # Second row of metrics for remaining metrics
        second_row_metrics = [
            "Best Trade", f"₹{best_trade:,.2f}" if not pd.isna(best_trade) else "N/A", "Worst Trade", f"₹{worst_trade:,.2f}" if not pd.isna(worst_trade) else "N/A",
            "Avg Win", f"₹{avg_win:,.2f}" if not pd.isna(avg_win) else "N/A", "Avg Loss", f"₹{avg_loss:,.2f}" if not pd.isna(avg_loss) else "N/A",
            "Start Capital", f"₹{starting_capital:,.2f}", "Final Capital", f"₹{metrics.final_capital:,.2f}", "", "", "", ""
        ]
        
        for col_idx, value in enumerate(second_row_metrics, 2):  # B=2 to O=15
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Color code based on performance
            if isinstance(value, str):
                if "₹" in str(value) and value != "N/A":
                    try:
                        numeric_val = float(str(value).replace("₹", "").replace(",", ""))
                        if numeric_val > 0:
                            cell.fill = positive_fill
                        elif numeric_val < 0:
                            cell.fill = negative_fill
                    except:
                        pass
                elif "%" in str(value):
                    try:
                        if float(str(value).replace("%", "")) > 50:
                            cell.fill = positive_fill
                        else:
                            cell.fill = neutral_fill
                    except:
                        pass
        current_row += 2

        # STRATEGY CONFIGURATION - Keep multi-line values visible with wrap and top align
        ws.merge_cells(f'A{current_row}:J{current_row}')
        config_header = ws[f'A{current_row}']
        config_header.value = "STRATEGY CONFIGURATION"
        config_header.font = header_font
        config_header.fill = header_fill
        config_header.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Get configuration and format properly
        additional_info_df = self._create_additional_info_table()
        
        # Create configuration in vertical layout
        for _, row in additional_info_df.iterrows():
            # Parameter name in column A-C
            ws.merge_cells(f'A{current_row}:C{current_row}')
            param_cell = ws[f'A{current_row}']
            param_cell.value = row['Key']
            param_cell.font = subheader_font
            param_cell.fill = summary_fill
            param_cell.border = border
            param_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Parameter value in column D-J, with proper formatting
            ws.merge_cells(f'D{current_row}:J{current_row}')
            value_cell = ws[f'D{current_row}']
            
            # Format multi-part values with line breaks for proper display
            value_text = str(row['Value'])
            if ',' in value_text and any(k in value_text.lower() for k in ['ema', 'fast', 'slow', 'enabled', 'activation']):
                # Format structured configuration data with proper line breaks
                parts = [p.strip() for p in value_text.split(',')]
                formatted_value = '\n'.join(f"• {p}" for p in parts)
            else:
                formatted_value = value_text
                
            value_cell.value = formatted_value
            value_cell.font = normal_font
            value_cell.border = border
            value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Calculate required height for multi-line content
            line_count = formatted_value.count('\n') + 1
            ws.row_dimensions[current_row].height = line_count * 18
            
            ws.row_dimensions[current_row].height = 25
            current_row += 1

        current_row += 1

        # DETAILED TRADES TABLE - Optimized for standard screen width
        ws.merge_cells(f'A{current_row}:J{current_row}')
        trades_header = ws[f'A{current_row}']
        trades_header.value = "DETAILED TRADES LOG"
        trades_header.font = header_font
        trades_header.fill = header_fill
        trades_header.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Trade headers - reduced columns to fit screen width
        trade_headers = ['#', 'Entry Time', 'Exit Time', 'Entry ₹', 'Exit ₹', 'Qty', 'Gross P&L', 'Net P&L', 'Exit Reason', 'Capital']
        
        for col_idx, header in enumerate(trade_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_row += 1

        # Add trade data with proper formatting
        for idx, (_, trade) in enumerate(trades_data.iterrows(), 1):
            trade_data = [
                idx,
                pd.to_datetime(trade['Entry Time']).strftime('%m/%d %H:%M') if pd.notna(trade['Entry Time']) else '',
                pd.to_datetime(trade['Exit Time']).strftime('%m/%d %H:%M') if pd.notna(trade['Exit Time']) else '',
                f"{trade['Entry Price']:.2f}" if pd.notna(trade['Entry Price']) else '',
                f"{trade['Exit Price']:.2f}" if pd.notna(trade['Exit Price']) else '',
                int(trade['Total Qty']) if pd.notna(trade['Total Qty']) else '',
                f"{trade['Gross P&L']:,.0f}" if pd.notna(trade['Gross P&L']) else '',
                f"{trade['Net P&L']:,.0f}" if pd.notna(trade['Net P&L']) else '',
                trade['Exit Reason'][:12] if pd.notna(trade['Exit Reason']) else '',
                f"{trade['Capital Outstanding']:,.0f}" if pd.notna(trade['Capital Outstanding']) else ''
            ]
            
            for col_idx, value in enumerate(trade_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Color code based on P&L
                if col_idx in [7, 8]:  # Gross P&L and Net P&L columns
                    if isinstance(value, str) and value.replace(',', '').replace('-', '').isdigit():
                        try:
                            numeric_val = float(value.replace(',', ''))
                            if numeric_val > 0:
                                cell.fill = positive_fill
                            elif numeric_val < 0:
                                cell.fill = negative_fill
                        except:
                            pass
            current_row += 1

        # Set optimal column widths to prevent text cutting and horizontal scrolling
        column_widths = {
            1: 4,   # Trade #
            2: 11,  # Entry Time
            3: 11,  # Exit Time  
            4: 9,   # Entry Price
            5: 9,   # Exit Price
            6: 6,   # Quantity
            7: 12,  # Gross P&L
            8: 12,  # Net P&L
            9: 13,  # Exit Reason
            10: 14  # Capital
        }
        
        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = width
        
        # Set row heights for better readability
        for row_num in range(1, current_row):
            if ws.row_dimensions[row_num].height is None:
                ws.row_dimensions[row_num].height = 20
        
        # Save the enhanced file with new naming convention
        timestamp = format_timestamp(datetime.now())
        # Get the Perplexity Combined results directory
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        results_dir = os.path.join(project_root, "results", "Back Test")
        os.makedirs(results_dir, exist_ok=True)
        
        enhanced_filename = os.path.join(results_dir, f"bt-{timestamp}-data.xlsx")
        wb.save(enhanced_filename)
        
        # Log the successful save
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Backtest results saved to: {enhanced_filename}")
        print(f"✅ Backtest results saved: {enhanced_filename}")
        
        return enhanced_filename

    def export_to_excel(self, output_dir: str = "results") -> str:
        """Export to Excel with gross P&L summary, config as a table, split trades and enhanced formatting."""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        additional_info_df = self._create_additional_info_table()

        # Compute gross P&L total
        gross_pnl_total = trades_df["Gross P&L"].dropna().replace("", 0).sum()

        # Config as wide table if possible
        if not additional_info_df.empty and "Key" in additional_info_df.columns and "Value" in additional_info_df.columns:
            config_wide_df = additional_info_df.set_index("Key").T.reset_index(drop=True)
        else:
            config_wide_df = pd.DataFrame([{"No Configuration": "Not Available"}])

        # Trades split
        trade_cols = trades_df.columns.tolist()
        break_idx = 7
        trades_df_1 = trades_df[trade_cols[:break_idx]]
        trades_df_2 = trades_df[trade_cols[break_idx:]]

        # Use new directory structure in Perplexity Combined folder
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # Go up to Perplexity Combined
        results_dir = os.path.join(base_dir, "results", "Back Test")
        os.makedirs(results_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(results_dir, f"bt-{timestamp}-data-detailed.xlsx")

        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            # Write config and trades tables starting from row 3 (Excel row 3, pandas startrow=2)
            config_wide_df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=2)
            trades_df_1.to_excel(writer, sheet_name="Sheet1", index=False, startrow=5)
            trades_df_2.to_excel(writer, sheet_name="Sheet1", index=False, startrow=5+len(trades_df_1)+2)

        from openpyxl import load_workbook  # type: ignore[reportMissingModuleSource]
        from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore[reportMissingModuleSource]
        from openpyxl.utils import get_column_letter  # type: ignore[reportMissingModuleSource]

        wb = load_workbook(excel_file)
        ws = wb.active

        # Set font for all cells and wrap text for string cells
        font16 = Font(size=18)
        wrap_align = Alignment(wrap_text=True, indent=1)
        for row in ws.iter_rows():
            for cell in row:
                # Text font size/wrap
                cell.font = font16
                cell.alignment = Alignment(
                    wrap_text=True,
                    indent=1,
                    horizontal="left" if cell.row > 1 else "center",
                    vertical="center"
                )
                # Make rows taller so text is readable
                try:
                    ws.row_dimensions[cell.row].height = max(ws.row_dimensions[cell.row].height or 0, 28)
                except:
                    ws.row_dimensions[cell.row].height = 28

        # Move Total Gross P&L to the top-left area (separate from table row) and make it prominent
        ws.merge_cells('A1:C2')
        ws.merge_cells('D1:F2')
        ws['A1'] = 'Total Gross P&L'
        ws['D1'] = gross_pnl_total

        # Style label (A1) — bigger and padded
        ws['A1'].font = Font(size=28, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, indent=2)
        ws['A1'].border = Border(
            top=Side(border_style="thick", color="000000"),
            bottom=Side(border_style="thick", color="000000"),
            left=Side(border_style="thick", color="000000"),
            right=Side(border_style="thick", color="000000"),
        )
        # Ensure rows 1-2 are tall enough for visual padding
        ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 40)
        ws.row_dimensions[2].height = max(ws.row_dimensions[2].height or 0, 40)

        # Style value (D1 merged) — very prominent
        ws['D1'].font = Font(size=45, bold=True)
        ws['D1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, indent=2)
        ws['D1'].number_format = '#,##0.00'
        # Assign a fresh Border built from A1's sides to avoid reusing a StyleProxy (unhashable)
        ws['D1'].border = Border(
            top=ws['A1'].border.top,
            bottom=ws['A1'].border.bottom,
            left=ws['A1'].border.left,
            right=ws['A1'].border.right,
        )
        if gross_pnl_total >= 0:
            ws['D1'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws['D1'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Ensure these columns are wider so the prominent cells have breathing room
        ws.column_dimensions[get_column_letter(1)].width = 18  # A
        ws.column_dimensions[get_column_letter(4)].width = 30  # D
        ws.column_dimensions[get_column_letter(5)].width = 30  # E
        ws.column_dimensions[get_column_letter(6)].width = 30  # F

        # Style config and trade headers: bold, center
        header_rows = [6, 6 + len(trades_df_1) + 2]
        for header_row in header_rows:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, size=16)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Column widths
        for col in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.value:
                    try:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                    except:
                        pass
            # increase padding and allow wider columns
            ws.column_dimensions[col_letter].width = min(max(max_length + 6, 20), 80)

        wb.save(excel_file)
        
        # Log the successful save
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Backtest detailed results saved to: {excel_file}")
        print(f"✅ Backtest detailed results saved: {excel_file}")
        
        # Also produce enhanced report automatically
        self.create_enhanced_excel_report(output_dir)
        return excel_file
BacktestResults = Results