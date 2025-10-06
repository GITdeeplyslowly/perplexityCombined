"""
backtest/results.py

Compute trading performance metrics for backtests, including:
- Total returns
- Win/loss stats
- Profit factor
- Equity curve and drawdown
- Trade durations and reason summaries
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass
from utils.time_utils import format_timestamp
import os
from openpyxl import load_workbook, Workbook  # type: ignore[reportMissingModuleSource]
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # type: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter  # type: ignore[reportMissingModuleSource]
from openpyxl.worksheet.page import PageMargins  # type: ignore[reportMissingModuleSource]
import logging

# Optional helpers (safe_divide and drawdown logic)
def safe_divide(numerator, denominator, default=0.0):
    try:
        if denominator == 0:
            return default
        return numerator / denominator
    except:
        return default

def calculate_drawdown(equity_curve: List[float]) -> float:
    max_drawdown = 0.0
    peak = equity_curve[0] if equity_curve else 0
    for value in equity_curve:
        if value > peak:
            peak = value
        drawdown = (peak - value) / peak if peak > 0 else 0
        max_drawdown = max(max_drawdown, drawdown)
    return max_drawdown * 100


@dataclass
class TradeResult:
    """Single structured record for a completed trade."""
    entry_time: datetime
    exit_time: datetime
    entry_price: float
    exit_price: float
    quantity: int
    pnl: float
    commission: float
    exit_reason: str


@dataclass
class TradingMetrics:
    """Aggregate statistics for evaluation."""
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    gross_profit: float = 0.0
    gross_loss: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    total_pnl: float = 0.0
    total_commission: float = 0.0
    net_pnl: float = 0.0
    best_trade: float = 0.0
    worst_trade: float = 0.0
    return_percent: float = 0.0
    final_capital: float = 0.0
    profit_factor: float = 0.0
    drawdown_percent: float = 0.0


class Results:
    """
    Result engine. Stores completed trades, tracks equity progression,
    and outputs performance metrics and CSV reports.
    """

    def __init__(self, initial_capital: float):
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.trades: List[TradeResult] = []
        self.equity_curve: List[Tuple[datetime, float]] = []
        self.config: Optional[Dict[str, Any]] = None  # Add config attribute

    def add_trade(self, trade_data: Dict[str, Any]) -> None:
        """Appends a trade and updates capital/equity."""
        trade = TradeResult(
            entry_time=trade_data['entry_time'],
            exit_time=trade_data['exit_time'],
            entry_price=trade_data['entry_price'],
            exit_price=trade_data['exit_price'],
            quantity=trade_data['quantity'],
            pnl=trade_data['pnl'],
            commission=trade_data['commission'],
            exit_reason=trade_data.get('exit_reason', ''),
        )
        self.trades.append(trade)
        self.current_capital += (trade.pnl - trade.commission)
        self.equity_curve.append((trade.exit_time, self.current_capital))

    def set_config(self, config: Dict[str, Any]):
        self.config = config

    def _create_additional_info_table(self) -> pd.DataFrame:
        """Create additional info table with indicators and parameters"""
        if not hasattr(self, 'config') or not self.config:
            return pd.DataFrame([{"Key": "Configuration", "Value": "Not Available"}])

        rows = []
        config = self.config
        strategy_config = config.get('strategy', config)  # fallback if not nested
        risk_config = config.get('risk', config)  # fallback if not nested

        # Indicators activated
        indicator_map = {
            'use_ema_crossover': 'EMA Crossover',
            'use_macd': 'MACD',
            'use_vwap': 'VWAP',
            'use_rsi_filter': 'RSI Filter',
            'use_htf_trend': 'HTF Trend',
            'use_bollinger_bands': 'Bollinger Bands',
            'use_stochastic': 'Stochastic',
            'use_atr': 'ATR'
        }
        active_indicators = [name for key, name in indicator_map.items()
                             if strategy_config.get(key, False)]
        rows.append({"Key": "Indicators Activated",
                     "Value": ", ".join(active_indicators) if active_indicators else "None"})

        # EMA parameters
        if strategy_config.get('use_ema_crossover', False):
            ema_params = f"Fast EMA: {strategy_config.get('fast_ema', 9)}, Slow EMA: {strategy_config.get('slow_ema', 21)}"
            rows.append({"Key": "EMA Parameters", "Value": ema_params})

        # MACD parameters
        if strategy_config.get('use_macd', False):
            macd_params = f"Fast: {strategy_config.get('macd_fast', 12)}, Slow: {strategy_config.get('macd_slow', 26)}, Signal: {strategy_config.get('macd_signal', 9)}"
            rows.append({"Key": "MACD Parameters", "Value": macd_params})

        # RSI parameters
        if strategy_config.get('use_rsi_filter', False):
            rsi_params = f"Period: {strategy_config.get('rsi_length', 14)}, Overbought: {strategy_config.get('rsi_overbought', 70)}, Oversold: {strategy_config.get('rsi_oversold', 30)}"
            rows.append({"Key": "RSI Parameters", "Value": rsi_params})

        # HTF Trend parameters
        if strategy_config.get('use_htf_trend', False):
            htf_params = f"HTF Period: {strategy_config.get('htf_period', 20)}"
            rows.append({"Key": "HTF Trend Parameters", "Value": htf_params})

        # SL points
        rows.append({"Key": "SL Points", "Value": str(risk_config.get('base_sl_points', 15))})

        # TP points
        tp_points = risk_config.get('tp_points', config.get('tp_points', [10, 25, 50, 100]))
        rows.append({"Key": "TP Points", "Value": ", ".join(map(str, tp_points))})

        # Trail SL info
        trail_enabled = risk_config.get('use_trail_stop', config.get('use_trail_stop', True))
        trail_activation = risk_config.get('trail_activation_points', config.get('trail_activation_points', 25))
        trail_distance = risk_config.get('trail_distance_points', config.get('trail_distance_points', 10))
        trail_info = f"Enabled: {trail_enabled}, Activation: {trail_activation} points, Distance: {trail_distance} points"
        rows.append({"Key": "Trailing Stop", "Value": trail_info})

        # --- LOGGING FOR DEBUG ---
        green_bars_req = strategy_config.get('consecutive_green_bars')
        logging.warning(f"DEBUG: consecutive_green_bars in strategy_config: {green_bars_req}")
        logging.warning(f"DEBUG: strategy_config keys: {list(strategy_config.keys())}")
        logging.warning(f"DEBUG: config keys: {list(config.keys())}")

        rows.append({"Key": "Green Bars Required for Entry", "Value": str(green_bars_req) })

        return pd.DataFrame(rows)

    def calculate_metrics(self) -> TradingMetrics:
        if not self.trades:
            return TradingMetrics(final_capital=self.initial_capital)

        total_trades = len(self.trades)
        wins = [t for t in self.trades if t.pnl > 0]
        losses = [t for t in self.trades if t.pnl < 0]

        gross_profit = sum(t.pnl for t in wins)
        gross_loss = sum(t.pnl for t in losses)
        total_commission = sum(t.commission for t in self.trades)
        total_pnl = gross_profit + gross_loss
        net_pnl = total_pnl - total_commission

        avg_win = safe_divide(gross_profit, len(wins))
        avg_loss = safe_divide(abs(gross_loss), len(losses))

        final_capital = self.initial_capital + net_pnl
        return_percent = safe_divide(net_pnl, self.initial_capital, 0.0) * 100
        drawdown = calculate_drawdown([v for _, v in self.equity_curve])

        return TradingMetrics(
            total_trades=total_trades,
            winning_trades=len(wins),
            losing_trades=len(losses),
            win_rate=100 * safe_divide(len(wins), total_trades),
            gross_profit=gross_profit,
            gross_loss=gross_loss,
            avg_win=avg_win,
            avg_loss=avg_loss,
            total_pnl=total_pnl,
            total_commission=total_commission,
            net_pnl=net_pnl,
            best_trade=max((t.pnl for t in self.trades), default=0.0),
            worst_trade=min((t.pnl for t in self.trades), default=0.0),
            return_percent=return_percent,
            final_capital=final_capital,
            profit_factor=safe_divide(gross_profit, abs(gross_loss), 0.0),
            drawdown_percent=drawdown,
        )

    def print_summary(self):
        """Show performance metrics in console."""
        m = self.calculate_metrics()
        print(f"\n{'='*60}")
        print("TRADING PERFORMANCE SUMMARY")
        print(f"{'='*60}")
        print(f"Total Trades        : {m.total_trades}")
        print(f"Win Rate (%)        : {m.win_rate:.2f}")
        print(f"Gross Profit        : â‚¹{m.gross_profit:.2f}")
        print(f"Gross Loss          : â‚¹{m.gross_loss:.2f}")
        print(f"Avg Win             : â‚¹{m.avg_win:.2f}")
        print(f"Avg Loss            : â‚¹{m.avg_loss:.2f}")
        print(f"Net P&L             : â‚¹{m.net_pnl:.2f}")
        print(f"Best Trade (P&L)    : â‚¹{m.best_trade:.2f}")
        print(f"Worst Trade (P&L)   : â‚¹{m.worst_trade:.2f}")
        print(f"Return (%)          : {m.return_percent:.2f}")
        print(f"Drawdown (%)        : {m.drawdown_percent:.2f}")
        print(f"Final Capital       : â‚¹{m.final_capital:,.2f}")
        print(f"Profit Factor       : {m.profit_factor:.2f}")
        print(f"Total Commission    : â‚¹{m.total_commission:.2f}")
        print(f"{'='*60}")

    def get_trade_summary(self) -> pd.DataFrame:
        rows = []
        capital = self.initial_capital

        # Insert starting capital row
        rows.append({
            "Entry Time": "",
            "Exit Time": "",
            "Entry Price": "",
            "Exit Price": "",
            "Lots": "",
            "Total Qty": "",
            "Gross P&L": "",
            "Commission": "",
            "Net P&L": "",
            "Exit Reason": "Starting Capital",
            "Duration (min)": "",
            "Capital Outstanding": round(capital, 2)
        })

        for t in self.trades:
            net_pnl = t.pnl - t.commission
            capital += net_pnl
            lots_display = getattr(t, 'lots_traded', t.quantity // getattr(t, 'lot_size', 1)) if hasattr(t, 'lot_size') else 'N/A'
            rows.append({
                "Entry Time": t.entry_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Exit Time": t.exit_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Entry Price": round(t.entry_price, 2),
                "Exit Price": round(t.exit_price, 2),
                "Lots": lots_display,
                "Total Qty": t.quantity,
                "Gross P&L": round(t.pnl, 2),
                "Commission": round(t.commission, 2),
                "Net P&L": round(net_pnl, 2),
                "Exit Reason": t.exit_reason,
                "Duration (min)": round((t.exit_time - t.entry_time).total_seconds() / 60, 2),
                "Capital Outstanding": round(capital, 2)
            })
        return pd.DataFrame(rows)

    def get_equity_curve(self) -> pd.DataFrame:
        """Return equity curve as DataFrame."""
        if not self.equity_curve:
            return pd.DataFrame()
        return pd.DataFrame(self.equity_curve, columns=["timestamp", "capital"])

    def export_to_csv(self, output_dir: str = "results") -> str:
        """Dump trades and performance to CSV. (No separate equity file)"""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        timestamp = format_timestamp(datetime.now())
        trades_file = os.path.join(output_dir, f"trades_{timestamp}.csv")

        # Create additional info table with trading configuration
        additional_info_df = self._create_additional_info_table()

        # Write both tables to the same CSV file
        with open(trades_file, 'w', newline='') as f:
            additional_info_df.to_csv(f, index=False)
            f.write('\n')  # Add empty line between tables
            trades_df.to_csv(f, index=False)

        return trades_file

    def create_enhanced_excel_report(self, output_dir: str = "results") -> str:
        """Create a comprehensive, professional Excel dashboard report with improved layout."""
        os.makedirs(output_dir, exist_ok=True)
        
        # Get data
        trades_df = self.get_trade_summary()
        metrics = self.calculate_metrics()
        
        # Remove starting capital row for trades processing
        trades_data = trades_df.iloc[1:].copy() if not trades_df.empty else trades_df.copy()
        starting_capital = trades_df.iloc[0]['Capital Outstanding'] if not trades_df.empty else self.initial_capital
        
        # Calculate statistics
        total_trades = len(trades_data)
        winning_trades = len(trades_data[trades_data['Gross P&L'] > 0]) if not trades_data.empty else 0
        losing_trades = len(trades_data[trades_data['Gross P&L'] < 0]) if not trades_data.empty else 0
        win_rate = (winning_trades / total_trades) * 100 if total_trades > 0 else 0
        
        best_trade = trades_data['Net P&L'].max() if not trades_data.empty else 0
        worst_trade = trades_data['Net P&L'].min() if not trades_data.empty else 0
        avg_win = trades_data[trades_data['Net P&L'] > 0]['Net P&L'].mean() if not trades_data.empty else 0
        avg_loss = trades_data[trades_data['Net P&L'] < 0]['Net P&L'].mean() if not trades_data.empty else 0
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Backtest Results"
        
        # Add buffer margins on all sides
        # Insert blank column A so no table uses column A (shifts everything right)
        ws.insert_cols(1)
        ws.column_dimensions['A'].width = 4
        # Top buffer: leave row 1 empty
        current_row = 2
        # Right buffer: empty column K
        ws.column_dimensions['K'].width = 4

        # Set page margins for better spacing
        ws.page_margins = PageMargins(left=1.0, right=0.7, top=0.75, bottom=0.75)

        # FONT SIZE: Increase all fonts by 1.5x
        title_font     = Font(size=27, bold=True, color="FFFFFF")  # 18 * 1.5
        header_font    = Font(size=23, bold=True, color="FFFFFF")  # 18 * 1.3 -> 23 (rounded)
        subheader_font = Font(size=16, bold=True)                   # 11 * 1.5
        normal_font    = Font(size=15)                              # 10 * 1.5

        title_fill = PatternFill(start_color="2E4BC6", end_color="2E4BC6", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        summary_fill = PatternFill(start_color="E8F1FF", end_color="E8F1FF", fill_type="solid")
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        # Metric-specific styling and layout
        metric_label_font = Font(size=14, bold=True)  # 11 * 1.3 -> 14 (rounded)
        metric_value_font = Font(size=16, bold=True)  # 12 * 1.3 -> 16 (rounded)
        metric_row_height = 40
        metric_col_widths = {col_idx: (14 if col_idx % 2 == 0 else 20) for col_idx in range(2, 12)}  # B..K

        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        # TITLE SECTION - keep tables out of column A, title starts at B
        ws.merge_cells(f'B{current_row}:K{current_row}')
        title_cell = ws[f'B{current_row}']
        title_cell.value = "BACKTEST RESULTS DASHBOARD"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 50
        current_row += 2

        # P&L Highlight - columns D to H (buffers at C/I)
        ws.merge_cells(f'D{current_row}:H{current_row}')
        ws[f'D{current_row}'].value = "TOTAL NET P&L"
        ws[f'D{current_row}'].font = Font(size=18, bold=True)
        ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'D{current_row}'].fill = summary_fill
        ws[f'D{current_row}'].border = border
        ws.row_dimensions[current_row].height = 60
        current_row += 1

        ws.merge_cells(f'D{current_row}:H{current_row}')
        ws[f'D{current_row}'].value = f"₹{metrics.net_pnl:,.2f}"
        ws[f'D{current_row}'].font = Font(size=26, bold=True)
        ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'D{current_row}'].border = border
        if metrics.net_pnl > 0:
            ws[f'D{current_row}'].fill = positive_fill
        else:
            ws[f'D{current_row}'].fill = negative_fill
        ws.row_dimensions[current_row].height = 80
        current_row += 2

        # PERFORMANCE SUMMARY - center block starting at column B and ending at column L
        ws.merge_cells(f'B{current_row}:L{current_row}')
        summary_cell = ws[f'B{current_row}']
        summary_cell.value = "PERFORMANCE SUMMARY"
        summary_cell.font = header_font
        summary_cell.fill = header_fill
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 40
        current_row += 1

        # Row 1: 4 pairs (8 columns B-I)
        row1_metrics = [
            "Total Trades", total_trades, "Win Rate", f"{win_rate:.2f}%", "Winning Trades", winning_trades, "Losing Trades", losing_trades
        ]
        for col_idx, value in enumerate(row1_metrics, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = metric_label_font if (col_idx % 2 == 0) else metric_value_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = metric_row_height
        current_row += 1

        # Row 2: 4 pairs (Gross/Commission/Net/Return)
        row2_metrics = [
            "Gross P&L", f"₹{metrics.total_pnl:,.2f}", "Commission", f"₹{metrics.total_commission:,.2f}", "Net P&L", f"₹{metrics.net_pnl:,.2f}", "Return %", f"{metrics.return_percent:.2f}%"
        ]
        for col_idx, value in enumerate(row2_metrics, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = metric_label_font if (col_idx % 2 == 0) else metric_value_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = metric_row_height
        current_row += 1

        # Row 3+: remaining metrics written in horizontal chunks that fit B..L (11 columns)
        # Note: Final Capital is handled separately (moved to columns D/E for prominence)
        row3_metrics = [
            "Best Trade", f"₹{best_trade:,.2f}" if not pd.isna(best_trade) else "N/A",
            "Worst Trade", f"₹{worst_trade:,.2f}" if not pd.isna(worst_trade) else "N/A",
            "Avg Win", f"₹{avg_win:,.2f}" if not pd.isna(avg_win) else "N/A",
            "Avg Loss", f"₹{avg_loss:,.2f}" if not pd.isna(avg_loss) else "N/A",
            "Start Capital", f"₹{starting_capital:,.2f}"
        ]
        # write in chunks so every table row occupies columns B..L only
        vals = row3_metrics
        i = 0
        chunk_width = 11  # B..L inclusive
        while i < len(vals):
            chunk = vals[i:i+chunk_width]
            for offset, value in enumerate(chunk, start=2):
                cell = ws.cell(row=current_row, column=offset, value=value)
                cell.font = metric_label_font if (offset % 2 == 0) else metric_value_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[current_row].height = metric_row_height
            current_row += 1
            i += chunk_width

        # Final Capital: place label/value at columns D (4) and E (5) and style prominently
        fc_label_cell = ws.cell(row=current_row, column=4, value="Final Capital")
        fc_value_cell = ws.cell(row=current_row, column=5, value=f"₹{metrics.final_capital:,.2f}")
        # Label styling: header-like
        fc_label_cell.font = header_font
        fc_label_cell.fill = header_fill
        fc_label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fc_label_cell.border = border
        # Value styling: prominent (larger white font on title_fill background)
        fc_value_cell.font = Font(size=26, bold=True, color="FFFFFF")  # 20 * 1.3 -> 26 (rounded)
        fc_value_cell.fill = title_fill
        fc_value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fc_value_cell.border = border
        ws.row_dimensions[current_row].height = metric_row_height
        current_row += 1

        # Ensure Final Capital columns are wide enough so value is fully visible
        try:
            ws.column_dimensions[get_column_letter(4)].width = max(ws.column_dimensions[get_column_letter(4)].width or 0, 30)
            ws.column_dimensions[get_column_letter(5)].width = max(ws.column_dimensions[get_column_letter(5)].width or 0, 40)
        except:
            pass

        # Set metric column widths starting at column B (B..L)
        metric_col_widths = {col_idx: (14 if col_idx % 2 == 0 else 20) for col_idx in range(2, 13)}  # 2..12 => B..L
        for col_idx, width in metric_col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # STRATEGY CONFIGURATION - start at column B so column A remains empty
        ws.merge_cells(f'B{current_row}:K{current_row}')
        config_header = ws[f'B{current_row}']
        config_header.value = "STRATEGY CONFIGURATION"
        config_header.font = header_font
        config_header.fill = header_fill
        config_header.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 45
        current_row += 1
         
        # Get configuration from the existing method (horizontal layout, compact two rows)
        additional_info_df = self._create_additional_info_table()
        
        # Create configuration in vertical layout
        for _, row in additional_info_df.iterrows():
            # Parameter name in columns B-D (keep A empty)
            ws.merge_cells(f'B{current_row}:D{current_row}')
            param_cell = ws[f'B{current_row}']
            param_cell.value = row['Key']
            param_cell.font = subheader_font
            param_cell.fill = summary_fill
            param_cell.border = border
            param_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Parameter value in columns E-K (kept away from column A)
            ws.merge_cells(f'E{current_row}:K{current_row}')
            value_cell = ws[f'E{current_row}']
            
            # Format multi-part values with line breaks for proper display
            value_text = str(row['Value'])
            if ',' in value_text and any(k in value_text.lower() for k in ['ema', 'fast', 'slow', 'enabled', 'activation']):
                # Format structured configuration data with proper line breaks
                parts = [p.strip() for p in value_text.split(',')]
                formatted_value = '\n'.join(f"• {p}" for p in parts)
            else:
                formatted_value = value_text
                
            value_cell.value = formatted_value
            value_cell.font = normal_font
            value_cell.border = border
            value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Calculate required height for multi-line content
            line_count = formatted_value.count('\n') + 1
            ws.row_dimensions[current_row].height = line_count * 18
            
            ws.row_dimensions[current_row].height = 25
            current_row += 1

        current_row += 1

        # DETAILED TRADES TABLE - start at column B so column A remains empty
        ws.merge_cells(f'B{current_row}:K{current_row}')
        trades_header = ws[f'B{current_row}']
        trades_header.value = "DETAILED TRADES LOG"
        trades_header.font = header_font
        trades_header.fill = header_fill
        trades_header.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Trade headers - reduced columns to fit screen width
        trade_headers = ['#', 'Entry Time', 'Exit Time', 'Entry ₹', 'Exit ₹', 'Qty', 'Gross P&L', 'Net P&L', 'Exit Reason', 'Capital']
        
        # place headers starting at column B
        for col_idx, header in enumerate(trade_headers, 2):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_row += 1

        # Add trade data with proper formatting
        # place trade rows starting at column B
        for idx, (_, trade) in enumerate(trades_data.iterrows(), 1):
            trade_data = [
                idx,
                pd.to_datetime(trade['Entry Time']).strftime('%m/%d %H:%M') if pd.notna(trade['Entry Time']) else '',
                pd.to_datetime(trade['Exit Time']).strftime('%m/%d %H:%M') if pd.notna(trade['Exit Time']) else '',
                f"{trade['Entry Price']:.2f}" if pd.notna(trade['Entry Price']) else '',
                f"{trade['Exit Price']:.2f}" if pd.notna(trade['Exit Price']) else '',
                int(trade['Total Qty']) if pd.notna(trade['Total Qty']) else '',
                f"{trade['Gross P&L']:,.0f}" if pd.notna(trade['Gross P&L']) else '',
                f"{trade['Net P&L']:,.0f}" if pd.notna(trade['Net P&L']) else '',
                trade['Exit Reason'][:12] if pd.notna(trade['Exit Reason']) else '',
                f"{trade['Capital Outstanding']:,.0f}" if pd.notna(trade['Capital Outstanding']) else ''
            ]
            
            # write trade data starting at column B (shift by +1)
            for col_offset, value in enumerate(trade_data, 2):
                cell = ws.cell(row=current_row, column=col_offset, value=value)
                # ensure wrap so long text does not overflow
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                
                # Color code based on P&L: Gross P&L at col 8, Net P&L at col 9 (with B start)
                if col_offset in (8, 9):
                    if isinstance(value, str) and value.replace(',', '').replace('-', '').replace(' ', '').isdigit():
                        try:
                            numeric_val = float(value.replace(',', ''))
                            if numeric_val > 0:
                                cell.fill = positive_fill
                            elif numeric_val < 0:
                                cell.fill = negative_fill
                        except:
                            pass
            current_row += 1

        # Set optimal column widths to prevent text cutting and horizontal scrolling
        # keys correspond to actual worksheet columns starting at B (2) -> K (11)
        column_widths = {
            2: 16,  # Trade #
            3: 20,  # Entry Time
            4: 20,  # Exit Time  
            5: 16,  # Entry Price
            6: 16,  # Exit Price
            7: 10,  # Quantity
            8: 18,  # Gross P&L
            9: 18,  # Net P&L
            10: 20, # Exit Reason
            11: 20  # Capital
        }
        # increase defaults so text fits; allow larger max
        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx)
            # increase min width to avoid clipping with larger fonts; trade columns given more room
            ws.column_dimensions[column_letter].width = max(width, 28)
         
        # Set row heights for better readability
        for row_num in range(1, current_row):
            if ws.row_dimensions[row_num].height is None:
                ws.row_dimensions[row_num].height = 20
        
        # Save the enhanced file
        timestamp = format_timestamp(datetime.now())
        enhanced_filename = os.path.join(output_dir, f"Enhanced_Backtest_Results_{timestamp}.xlsx")
        wb.save(enhanced_filename)
        
        return enhanced_filename

    def export_to_excel(self, output_dir: str = "results") -> str:
        """Export to Excel with gross P&L summary, config as a table, split trades and enhanced formatting."""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        additional_info_df = self._create_additional_info_table()

        # Compute gross P&L total
        gross_pnl_total = trades_df["Gross P&L"].dropna().replace("", 0).sum()

        # Config as wide table if possible
        if not additional_info_df.empty and "Key" in additional_info_df.columns and "Value" in additional_info_df.columns:
            config_wide_df = additional_info_df.set_index("Key").T.reset_index(drop=True)
        else:
            config_wide_df = pd.DataFrame([{"No Configuration": "Not Available"}])

        # Trades split
        trade_cols = trades_df.columns.tolist()
        break_idx = 7
        trades_df_1 = trades_df[trade_cols[:break_idx]]
        trades_df_2 = trades_df[trade_cols[break_idx:]]

        timestamp = format_timestamp(datetime.now())
        excel_file = os.path.join(output_dir, f"trades_{timestamp}.xlsx")

        # Create workbook and worksheet manually so we control column ranges (B..K used for tables; L reserved blank)
        wb = Workbook()
        ws = wb.active

        # Ensure column A is empty (buffer) and column L is reserved as empty buffer
        try:
            ws.insert_cols(1)
        except:
            pass
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['L'].width = 3

        # Helper: write dataframe in horizontal chunks that fit within B..K (10 columns)
        max_table_cols = 10  # B..K inclusive
        write_start_row = 3

        # small inline helper (no new function defs) - chunk columns and write each chunk as a separate table block
        df_list = [("CONFIG", config_wide_df), ("TRADES1", trades_df_1), ("TRADES2", trades_df_2)]
        for name, df in df_list:
            if df is None or df.empty:
                continue
            cols = list(df.columns)
            col_ptr = 0
            while col_ptr < len(cols):
                chunk = cols[col_ptr:col_ptr + max_table_cols]
                # header row
                for j, cname in enumerate(chunk, start=2):  # start at column B
                    ws.cell(row=write_start_row, column=j, value=cname)
                # data rows
                for r_idx, row in enumerate(df[chunk].itertuples(index=False, name=None), start=1):
                    for j, val in enumerate(row, start=2):
                        ws.cell(row=write_start_row + r_idx, column=j, value=val)
                # advance pointer and row pointer (leave one blank row between chunked tables)
                rows_used = len(df.index) + 1
                write_start_row += rows_used + 1
                col_ptr += max_table_cols

        # Apply page margins and base styling
        ws.page_margins = PageMargins(left=1.0, right=1.0, top=0.9, bottom=0.9)
        base_font = Font(size=14)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.font = base_font
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        # Move Total Gross P&L to the top-left area (A1..F2 region), keep column L empty
        ws.merge_cells('A1:C2')
        ws.merge_cells('D1:F2')
        ws['A1'] = 'Total Gross P&L'
        ws['D1'] = gross_pnl_total
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['A1'].border = Border(top=Side(border_style="thick", color="000000"),
                                 bottom=Side(border_style="thick", color="000000"),
                                 left=Side(border_style="thick", color="000000"),
                                 right=Side(border_style="thick", color="000000"))
        ws['D1'].font = Font(size=24, bold=True)
        ws['D1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['D1'].number_format = '#,##0.00'
        # Avoid assigning a StyleProxy directly (unhashable). Create a fresh Border using A1's side attributes.
        a1_border = ws['A1'].border
        new_border = Border(
            top=Side(border_style=a1_border.top.border_style, color=a1_border.top.color),
            bottom=Side(border_style=a1_border.bottom.border_style, color=a1_border.bottom.color),
            left=Side(border_style=a1_border.left.border_style, color=a1_border.left.color),
            right=Side(border_style=a1_border.right.border_style, color=a1_border.right.color),
        )
        ws['D1'].border = new_border
        ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 40)
        ws.row_dimensions[2].height = max(ws.row_dimensions[2].height or 0, 40)

        # Ensure column L remains empty and narrow (buffer)
        ws.column_dimensions['L'].width = 3

        # Set reasonable column width caps to avoid horizontal scrolling; tables limited to B..K
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            current = ws.column_dimensions[letter].width or 10
            ws.column_dimensions[letter].width = min(max(int(current), 12), 32)

        wb.save(excel_file)
        # Also produce enhanced report automatically
        self.create_enhanced_excel_report(output_dir)
        return excel_file
BacktestResults = Results